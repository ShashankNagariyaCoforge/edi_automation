
import sys
import openpyxl
from pathlib import Path

def inspect(file_path):
    print(f"Inspecting: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet in wb.sheetnames:
        print(f"\nSheet: {sheet}")
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
            valid_cells = [str(c)[:50] for c in row if c is not None]
            if valid_cells:
                print(f"Row {i}: {valid_cells}")

if __name__ == "__main__":
    inspect(sys.argv[1])
