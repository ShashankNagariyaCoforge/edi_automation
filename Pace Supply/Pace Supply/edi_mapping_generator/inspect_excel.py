import openpyxl
from pathlib import Path

def inspect_file():
    path = Path("/home/azureuser/Documents/projects/edi_automation/edi_mapping_generator/input/inbound_X12_to_oracle.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    
    # Check all sheets
    for sheet_name in wb.sheetnames:
        print(f"--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        
        # Scan first 50 rows
        for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True), start=1):
            # Convert to string to make search easier
            row_str = [str(cell) if cell is not None else "" for cell in row]
            
            # Print non-empty rows
            if any(row_str):
                print(f"Row {i}: {row_str[:15]}") # Print first 15 cols
                
    wb.close()

if __name__ == "__main__":
    inspect_file()
