"""
Excel reader module.
Phase 1: Read structure/layout to allow deterministic updates later.
"""
from typing import Dict, List, Any
from pathlib import Path
from openpyxl import load_workbook
from .logger import get_logger


def read_erp_structure(file_path: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Phase 1: Read the ERP Definition Excel to extract the structure.
    Returns a dictionary grouping rows by Record Number.
    """
    logger = get_logger()
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    logger.info(f"Reading ERP structure from: {path.name}")
    
    wb = load_workbook(path, data_only=True)
    
    target_sheet = None
    for name in wb.sheetnames:
        if "inbound" in name.lower() and "oracle" in name.lower():
            target_sheet = wb[name]
            break
    
    if not target_sheet:
        target_sheet = wb.active
        logger.info(f"Using active sheet: {target_sheet.title}")

    structure = {}
    current_record = None
    
    processed_count = 0
    # iter_rows with values_only=True yields tuples of cell values
    for i, row in enumerate(target_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
            
        field_name = str(row[0]).strip() if row[0] else ""
        
        # 1. Check for Record Header in Column A
        # Example: "0010 Record" or "1000 Record"
        if "Record" in field_name:
            parts = field_name.split()
            # Check if first part is digits (e.g. "0010")
            if parts and parts[0].isdigit():
                current_record = parts[0].zfill(4)
                logger.debug(f"Row {i}: Found new record section header: {current_record}")
                # We usually don't map the header row itself, but we should continue to next row
                continue
                
        # 2. Extract specific record ref from Column F (index 5)
        record_ref = None
        if len(row) > 5 and row[5]:
             record_ref = row[5]

        rec_id = None
        if record_ref:
            try:
                val = float(record_ref)
                rec_id = str(int(val))
            except:
                rec_id = str(record_ref).strip()
            
            if rec_id.isdigit() and len(rec_id) <= 4:
                rec_id = rec_id.zfill(4)
        
        # 3. Fallback to current section
        if not rec_id:
            rec_id = current_record
            
        # If still no record ID or no field name, skip
        if not rec_id:
            # logger.debug(f"Row {i}: Skipping, no record ID. Field: {field_name}")
            continue
            
        if not field_name:
             continue

        # Valid row to map
        item = {
            "row_idx": i,
            "field_name": field_name,
            "record_ref": rec_id
        }
        
        if rec_id not in structure:
            structure[rec_id] = []
        structure[rec_id].append(item)
        processed_count += 1
        
    logger.info(f"Read structure: {len(structure)} record types, {processed_count} fields.")
    wb.close()
    return structure
