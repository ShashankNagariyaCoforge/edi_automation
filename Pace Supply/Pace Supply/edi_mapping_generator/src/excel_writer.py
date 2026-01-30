"""
Excel writer module for generating output mapping files.
Preserves original Excel formatting by copying and editing the template.
"""
import shutil
import os
import pandas as pd
from dotenv import load_dotenv
from pathlib import Path

# Load environment variables
load_dotenv(Path(__file__).parent.parent / ".env")
from typing import Dict, Any, List
from datetime import datetime
from openpyxl import load_workbook
from logger import get_logger


def write_mapping_output(
    structure: Dict[str, List[Dict[str, Any]]],
    mappings: Dict[str, Dict[str, Dict[str, Any]]],
    output_path: str
) -> str:
    """
    Phase 4: Deterministic write to Excel using pre-read structure.
    Populates Columns B, C, D, E, J.
    """
    logger = get_logger()
    
    # 1. Setup paths
    output_dir = Path(output_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"generated_mapping_{timestamp}.xlsx"
    
    # Template: Source from configured path
    input_template_path = None
    
    # Check env var first
    env_erp_path = os.getenv("ERP_DEFINITION_PATH")
    base_dir = Path(__file__).parent.parent
    
    erp_files = []
    if env_erp_path:
            p = Path(env_erp_path)
            if not p.is_absolute():
                p = base_dir / p
            erp_files.append(p)

    # Fallbacks
    erp_files.append(base_dir / "input" / "inbound_X12_to_oracle.xlsx")
    erp_files.append(base_dir / "inbound_X12_to_oracle.xlsx")
    erp_files.append(base_dir / "ERP_Definition.xlsx")
    
    for f in erp_files:
        if f.exists():
            input_template_path = f
            break
            
    if not input_template_path:
         raise FileNotFoundError(f"Mapping template file not found. Searched in: {[str(f) for f in erp_files]}")

    logger.info(f"Copying template {input_template_path} to: {output_file}")
    shutil.copy2(input_template_path, output_file)
    
    # 2. Load
    wb = load_workbook(output_file)
    # Active sheet or find target
    ws = wb.active
    for name in wb.sheetnames:
        if "inbound" in name.lower() and "oracle" in name.lower():
            ws = wb[name]
            break

    logger.info(f"Writing to sheet: '{ws.title}' (Phase 4)")
    
    # 3. Populate
    updated_count = 0
    
    # Col Mapping: 1-based indices
    # B=2, C=3, D=4, E=5, J=10
    col_map = {"B": 2, "C": 3, "D": 4, "E": 5}
    
    for rec_id, fields_list in structure.items():
        rec_mappings = mappings.get(rec_id, {})
        
        for field_item in fields_list:
            row_idx = field_item["row_idx"]
            f_name = field_item["field_name"]
            
            # Get LLM/System result
            # Try to match loose? No, Phase 3 should have returned keys matching inputs.
            val_map = rec_mappings.get(f_name)
            
            if not val_map:
                continue
                
            # Write columns
            for key, col_idx in col_map.items():
                val = val_map.get(key)
                if val is not None:
                     ws.cell(row=row_idx, column=col_idx).value = val
            
            updated_count += 1
            
    logger.info(f"Updated {updated_count} fields in Excel.")
    
    # 4. Save
    wb.save(output_file)
    return str(output_file)



def create_summary_sheet(
    output_file: str,
    mappings: Dict[str, Dict[str, Dict[str, Any]]],
    stats: Dict[str, Any]
):
    """
    Add a summary sheet to the existing output Excel file using openpyxl.
    """
    logger = get_logger()
    
    try:
        wb = load_workbook(output_file)
        
        if "Summary" in wb.sheetnames:
            del wb["Summary"]
            
        ws = wb.create_sheet("Summary")
        
        ws['A1'] = "Processing Summary"
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
        
        metrics = [
            ("Total Records Processed", stats.get("total_records", 0)),
            ("Total Fields Mapped", stats.get("total_fields", 0)),
            ("Processing Time (seconds)", f"{stats.get('processing_time', 0):.2f}"),
            ("Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ]
        
        for idx, (metric, value) in enumerate(metrics, start=3):
            ws[f'A{idx}'] = metric
            ws[f'B{idx}'] = value
            
        wb.save(output_file)
        logger.info("Summary sheet added")
        
    except Exception as e:
        logger.error(f"Failed to add summary sheet: {e}")
