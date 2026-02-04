from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, Any, List
import shutil
import time

class ExcelBuilder856:
    """
    Builds the 856 Output Excel file.
    Follows format of 'PaceSupply_856_Outbound.xlsx'.
    """
    
    def __init__(self, template_path: str = "856/PaceSupply_856_Outbound.xlsx"):
        # Resolve path relative to project root (parent of src)
        # __file__ is src/flow_856/excel_builder.py
        # parent.parent is src/
        # parent.parent.parent is project root
        project_root = Path(__file__).parent.parent.parent
        self.template_path = project_root / template_path
        if not self.template_path.exists():
            # Fallback path logic if needed
            pass
            
    def build_excel(self, mappings: Dict[str, Any], output_path: str) -> str:
        """
        Populate the template with mappings.
        mappings structure: 
        { 
          "mappings": [ 
             { "segment": "BSN", "element": "BSN01", "erp_record": "0010", "erp_field": "DOC...", "logic": "..." } 
          ]
        }
        """
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        final_path = Path(output_path) / f"856_Mapping_{timestamp}.xlsx"
        
        # Copy template
        shutil.copy(self.template_path, final_path)
        
        wb = load_workbook(final_path)
        if "ANSI X12 Mapping" in wb.sheetnames:
            ws = wb["ANSI X12 Mapping"]
        else:
            ws = wb.active
            
        # The user said: "output file should have exactly those columns as in this sheet"
        # "input file PaceSupply_856_Outbound.xlsx in sheet 'ANSI X12 Mapping'"
        # Columns:
        # A: Field Name (Segment + Element + Description)
        # B, C, D... 
        # E: Mapping (Record/Position)
        
        # We need to find rows corresponding to the Segments we found.
        # But wait, the Template likely already HAS rows for segments.
        # OR we need to Append?
        # The user said: "wherever we have mandatory mentioned in segments all of those segments we need to put in mapping output file."
        # This implies we might need to CLEAR the sheet and write fresh, OR append.
        # But if we use it as a template, maybe it has a header?
        # Let's assume we append to the end of the sheet or overwrite if logic permits.
        
        # Strategy:
        # 1. Clear existing data rows (keep header).
        # 2. Write each mandatory segment/field.
        
        # Header Row is 1. Start data at 2.
        # Template Columns:
        # A (1): Seg.
        # B (2): Occ./Max
        # C (3): Element
        # D (4): Type ('Source' or 'Constant')
        # E (5): Source (Mapping)
        # F (6): Hardcode
        # G (7): Meaning (Description)
        # H (8): Optional/Mandatory
        
        start_row = 2
        mapped_data = mappings.get("mappings", [])
        
        last_segment = None
        
        for i, item in enumerate(mapped_data):
            row_idx = start_row + i
            
            # 1. Segment (Col A) - Grouping logic
            segment = item.get("segment", "")
            if segment == last_segment:
                self._safe_write(ws, row_idx, 1, None)
            else:
                self._safe_write(ws, row_idx, 1, segment)
                last_segment = segment
                
            # 3. Element (Col C)
            element = item.get("element", "")
            self._safe_write(ws, row_idx, 3, element)
            
            # 6. Description (Col G) - Default
            desc = item.get("logic", "")
            self._safe_write(ws, row_idx, 7, desc)
            
            # 7. Requirement (Col H) - Default Mandatory
            self._safe_write(ws, row_idx, 8, "Mandatory")
            
            # MAPPING LOGIC
            typ = item.get("type", "")
            hardcode = item.get("hardcode", "")
            
            rec = item.get("erp_record", "")
            field = item.get("erp_field", "")
            pos = item.get("erp_position", "")

            if typ:
                self._safe_write(ws, row_idx, 4, typ)
                
                if typ == "Source":
                     if rec and pos:
                         self._safe_write(ws, row_idx, 5, f"{rec}/{pos}")
                     self._safe_write(ws, row_idx, 7, field)
                     
                elif typ in ["Constant", "Translation"]:
                     self._safe_write(ws, row_idx, 6, hardcode)
                     self._safe_write(ws, row_idx, 7, desc)
                     
                     if typ == "Translation" and rec and pos:
                         self._safe_write(ws, row_idx, 5, f"{rec}/{pos}")
                else:
                     self._safe_write(ws, row_idx, 7, desc)
            else:
                self._safe_write(ws, row_idx, 7, desc)
                if rec and field:
                    self._safe_write(ws, row_idx, 4, "Source")
                    self._safe_write(ws, row_idx, 7, field)
                    if pos:
                         self._safe_write(ws, row_idx, 5, f"{rec}/{pos}")

        wb.save(final_path)
        return str(final_path)

    def _safe_write(self, ws, row, col, value):
        from openpyxl.cell.cell import MergedCell
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    top_left.value = value
                    return
        else:
            cell.value = value

