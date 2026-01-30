from typing import Dict, Any, List
from pathlib import Path
import json
from openpyxl import load_workbook
from src.ai_client import AIClient
from src.logger import get_logger

class MappingEngine856:
    """
    Engine for generating 856 mappings.
    """
    def __init__(self, ai_client: AIClient):
        self.ai_client = ai_client
        self.logger = get_logger()
        self.definitions = []

    def load_definitions(self, file_path: str):
        """
        Load ERP Definitions from Excel.
        Expected structure: Col A=Field Name, Col B=Record Number.
        """
        self.logger.info(f"Loading ERP definitions from: {file_path}")
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        self.definitions = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]: continue
            
            # Format: Name, Record, Position, Length, Type...
            field_name = str(row[0]).strip()
            record_num = str(row[1]).strip() if len(row) > 1 and row[1] else "0000"
            position = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            
            # Normalize record num
            if record_num.isdigit():
                record_num = record_num.zfill(4)
                
            self.definitions.append({
                "field_name": field_name,
                "record_num": record_num,
                "position": position,
                "row_idx": i
            })
            
        self.logger.info(f"Loaded {len(self.definitions)} ERP fields")

    def generate_mapping(self, mandatory_segments: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Generate mappings for mandatory segments using LLM.
        """
        self.logger.info("Generating 856 mappings via LLM...")
        
        # Prepare context for LLM
        erp_list_str = "\n".join([f"{d['record_num']} - {d['field_name']}" for d in self.definitions])
        
        prompt = self._build_mapping_prompt(mandatory_segments, erp_list_str)
        
        try:
            response = self.ai_client.get_completion(
                prompt,
                system_prompt="You are an Expert EDI Mapper. Map Vendor Spec fields to Oracle ERP Interface Tables."
            )
            
            mapping_result = self._parse_json(response)
            
            mapping_result = self._parse_json(response)
            
            # Post-process to add Position and Validation
            raw_mappings = mapping_result.get("mappings", [])
            final_mappings = []
            
            # Index mandatory segments for lookups
            # Map "Seg+Elem" -> Field Dict
            indexed_pdf_fields = {}
            for seg in mandatory_segments:
                s_code = seg.get("segment")
                for f in seg.get("fields", []):
                    f_id = f.get("id") # e.g. ST01
                    # Normalize ID. Sometimes it's ST01, sometimes just 01? 
                    # Usually extract_mandatory returns full ID "ST01"
                    indexed_pdf_fields[f_id] = f

            for m in raw_mappings:
                target_field = m.get("erp_field")
                elem_id = m.get("element") # e.g. ST01
                
                # Retrieve original PDF constraints
                pdf_field = indexed_pdf_fields.get(elem_id, {})
                values = pdf_field.get("values", [])
                description = pdf_field.get("description", "").lower() + " " + m.get("logic", "").lower()
                
                # Initialize Extended Attributes
                typ = "Source" 
                hardcode = ""
                
                # Rule 1: Codes Logic (Constant vs Translation)
                if values:
                    if len(values) == 1:
                        typ = "Constant"
                        hardcode = values[0]
                    elif len(values) > 1:
                        typ = "Translation"
                        hardcode = ", ".join(values)
                
                # Rule 2: Description Logic if not Constant/Translation (or override?)
                # User says: "if no code and if you read description..."
                if not values:
                    if "sequence" in description:
                        typ = "Sequence"
                    elif "subordinate" in description and "hierarchical" in description:
                        # HL02 specific
                        typ = "Inherit"
                    elif "number of" in description or "count" in description:
                        typ = "Count"
                    else:
                        # Fallback
                        if not target_field:
                            typ = "" # Unknown/Manual
                        else:
                            typ = "Source"

                m["type"] = typ
                m["hardcode"] = hardcode
                m["description"] = description # Persist for Service/Builder usage

                if target_field:
                    # Lookup position
                    match = next((d for d in self.definitions if d["field_name"] == target_field), None)
                    if match:
                        m["erp_position"] = match["position"]
                
                final_mappings.append(m)
                        
            mapping_result["mappings"] = final_mappings
            return mapping_result
            
        except Exception as e:
            self.logger.error(f"Mapping generation failed: {e}")
            return {}

    def _build_mapping_prompt(self, segments: List[Dict], erp_fields: str) -> str:
        return f"""
You have a list of MANDATORY Segments/Fields from a Vendor 856 Spec.
You have a list of Oracle ERP Interface Fields (Record Number - Field Name).

YOUR TASK:
For each Vendor Field, find the BEST MATCH in the ERP Interface Fields.

OUTPUT FORMAT (JSON):
{{
  "mappings": [
    {{
       "segment": "BSN",
       "element": "BSN01",
       "erp_record": "0010",
       "erp_field": "TP_LOCATION_CODE_EXT",
       "logic": "Direct Map"
    }},
    ...
  ]
}}

INPUT DATA:

### VENDOR SPEC (Mandatory Only):
{json.dumps(segments, indent=2)}

### ERP INTERFACE FIELDS (Available Targets):
{erp_fields}

INSTRUCTIONS:
1. Map strictly based on standard EDI knowledge.
2. Example: BSN01 (Trans Set Purpose) -> often maps to a Control or Attribute field if no direct match, or specialized instruction.
3. If uncertain, map to closest semantic match or leave erp_field empty.
4. The goal is to fill the 'ANSI X12 Mapping' Excel sheet later.
"""

    def _parse_json(self, response: str):
        import json
        try:
            if "```json" in response:
                response = response.split("```json")[-1].split("```")[0].strip()
            elif "```" in response:
                response = response.strip("`")
            return json.loads(response)
        except:
            return {}
