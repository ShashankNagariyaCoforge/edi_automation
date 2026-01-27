import sys
from pathlib import Path
from openpyxl import load_workbook

def verify_record_0020(filepath):
    print(f"Verifying Record 0020 in: {filepath}")
    wb = load_workbook(filepath, data_only=True)
    ws = wb[' Inbound X12 to Oracle']
    
    count_0020 = 0
    failures_0020 = 0
    
    print("-" * 60)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        
        record_ref = str(row[5]) if row[5] else "" # Column F
        if "0020" in record_ref:
            count_0020 += 1
            field_name = row[0]
            logic = row[9] # Column J
            
            status = "OK"
            if logic and "Cannot determine mapping" in str(logic):
                status = "FAIL"
                failures_0020 += 1
                
            print(f"0020 | {field_name:<40} | {status} | {'Logic: ' + str(logic)[:50] if logic else ''}")

    print("-" * 60)
    print(f"Total 0020 Fields: {count_0020}")
    print(f"Failures: {failures_0020}")
    
    if failures_0020 == 0 and count_0020 > 0:
        print("SUCCESS: Record 0020 is fully mapped.")
    else:
        print("FAILURE: Record 0020 still has issues.")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        verify_record_0020(sys.argv[1])
