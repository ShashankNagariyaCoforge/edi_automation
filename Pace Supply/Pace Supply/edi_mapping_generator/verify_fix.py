import sys
from pathlib import Path
from openpyxl import load_workbook

def verify_output(filepath):
    print(f"Verifying: {filepath}")
    wb = load_workbook(filepath, data_only=True)
    ws = wb[' Inbound X12 to Oracle']
    
    errors_found = 0
    checked_records = set()
    
    print(f"\nScanning for 'Cannot determine mapping' errors...")
    print("-" * 60)

    # Column J is index 9 (0-based) or 'J'
    # Iterate rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        
        field_name = row[0]
        record_ref = row[5] # Column F
        logic = row[9]      # Column J
        
        if not field_name or not record_ref:
            continue
            
        record_ref = str(record_ref).split('.')[0].zfill(4)
        checked_records.add(record_ref)

        if logic and "Cannot determine mapping" in str(logic):
            # Only flag if it's the "not in KB" generic error, specific reasons might be valid
            if "no Knowledge Base definition found" in str(logic) or "not in KB" in str(logic):
                print(f"[FAIL] Record {record_ref} | Field: {field_name} | Logic: {logic}")
                errors_found += 1
    
    print("-" * 60)
    print(f"Checked {len(checked_records)} record types.")
    
    if errors_found == 0:
        print("SUCCESS: No unmapped fields due to missing definitions found!")
    else:
        print(f"FAILURE: Found {errors_found} unmapped fields.")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        verify_output(sys.argv[1])
    else:
        print("Please provide output file path")
